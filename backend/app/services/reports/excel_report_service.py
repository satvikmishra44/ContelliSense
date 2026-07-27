import os
from typing import List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.chart import LineChart, Reference
from sqlalchemy.orm import Session

from app.db.models.analysis import Analysis
from app.db.models.recommendation import Recommendation
from app.db.models.video import Video


class ExcelReportService:
    """
    Generates a multi-sheet Excel report with:
      - Summary dashboard
      - Recommendations sheet
      - Historical video data
    """

    def __init__(self, db: Session) -> None:
        self.db = db

    def generate_report(self, analysis_uuid: str) -> str:
        analysis = (
            self.db.query(Analysis)
            .filter(Analysis.analysis_uuid == analysis_uuid)
            .first()
        )
        if not analysis:
            raise ValueError("Analysis not found")

        channel_videos: List[Video] = (
            self.db.query(Video)
            .filter(Video.channel_id == analysis.channel_id)
            .limit(200)
            .all()
        )
        recs: List[Recommendation] = (
            self.db.query(Recommendation)
            .filter(Recommendation.analysis_id == analysis.id)
            .all()
        )

        wb = Workbook()

        # Summary sheet
        summary_ws = wb.active
        summary_ws.title = "Summary"
        summary_ws["A1"] = "YouTube Content Strategy Report"
        summary_ws["A1"].font = Font(size=16, bold=True)
        summary_ws["A3"] = "Analysis UUID"
        summary_ws["B3"] = analysis.analysis_uuid

        # Recommendations sheet
        rec_ws = wb.create_sheet("Recommendations")
        headers = [
            "Title",
            "Hook",
            "Thumbnail Idea",
            "Summary",
            "Target Audience",
            "Why It Works",
            "Evidence",
            "Trend",
            "Risk",
            "Effort",
            "Expected CTR",
            "Search Potential",
            "Virality",
            "Confidence",
            "Hit Probability",
            "Publishing Window",
        ]
        for col, header in enumerate(headers, start=1):
            cell = rec_ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="CCCCCC")

        for row_idx, rec in enumerate(recs, start=2):
            rec_ws.cell(row=row_idx, column=1, value=rec.title)
            rec_ws.cell(row=row_idx, column=2, value=rec.hook)
            rec_ws.cell(row=row_idx, column=3, value=rec.thumbnail_idea)
            rec_ws.cell(row=row_idx, column=4, value=rec.summary)
            rec_ws.cell(row=row_idx, column=5, value=rec.target_audience)
            rec_ws.cell(row=row_idx, column=6, value=rec.why_it_should_work)
            rec_ws.cell(row=row_idx, column=7, value=rec.supporting_evidence)
            rec_ws.cell(row=row_idx, column=8, value=rec.trend_explanation)
            rec_ws.cell(row=row_idx, column=9, value=rec.risk_factors)
            rec_ws.cell(row=row_idx, column=10, value=rec.estimated_effort)
            rec_ws.cell(row=row_idx, column=11, value=rec.expected_ctr)
            rec_ws.cell(row=row_idx, column=12, value=rec.search_potential)
            rec_ws.cell(row=row_idx, column=13, value=rec.virality_score)
            rec_ws.cell(row=row_idx, column=14, value=rec.confidence_score)
            rec_ws.cell(row=row_idx, column=15, value=rec.hit_probability)
            rec_ws.cell(row=row_idx, column=16, value=rec.publishing_window)

        # Historical sheet with simple chart
        hist_ws = wb.create_sheet("History")
        hist_ws["A1"] = "Video Title"
        hist_ws["B1"] = "Views"
        hist_ws["C1"] = "Engagement Rate"

        for idx, v in enumerate(channel_videos, start=2):
            hist_ws.cell(row=idx, column=1, value=v.title)
            hist_ws.cell(row=idx, column=2, value=v.views)
            hist_ws.cell(row=idx, column=3, value=v.engagement_rate)

        chart = LineChart()
        chart.title = "Views Over Videos"
        data = Reference(hist_ws, min_col=2, min_row=1, max_row=len(channel_videos) + 1)
        chart.add_data(data, titles_from_data=True)
        hist_ws.add_chart(chart, "E2")

        # Output path
        os.makedirs("reports", exist_ok=True)
        filename = f"reports/analysis_{analysis_uuid}.xlsx"
        wb.save(filename)

        analysis.report_path = filename
        self.db.commit()

        return filename